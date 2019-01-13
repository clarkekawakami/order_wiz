from openpyxl import load_workbook
from dicttoxml import dicttoxml
from xml.dom.minidom import parseString
from pathlib import Path
import PIL
import glob

print('*****************************************')
print('**                                     **')
print('** Welcome to the F19 DPS Order Wizard **')
print('**                                     **')

# loop until user says you are done
DONE = False

while DONE == False:
  # list xlsx files currently in the inputs folder
  print('*****************************************')
  print(" ")
  print("Files available to process: ")
  print(" ")

  available_files = glob.glob("inputs/*.xlsx")

  for xlsx in available_files:
    print(xlsx[7:])

  print(" ")

  # ask user for input filename... file is expected in inputs/ subdirectory
  print('*****************************************')
  print(" ")
  input_fn = input("Enter filename to process: ")
  my_input = Path("inputs/" + input_fn)
  


  # if input file exists then off to the races
  if my_input.is_file():

    print(" ")
    print(input_fn + " is a:")
    print("  #1: SKI PRSN order")
    print("  #2: PHANTOM PRSN order")
    print("  #3: DREAMTIME order")
    print("  #4: DEMO/RENTAL order")
    print(" ")
    input_ordertype = input("Enter order type [1, 2, 3 or 4]: ")

    # my_input = Path("inputs/" + input_fn)

    # ship date constants... change as required... date format may have to change to work with IQMS ??
    dt_ship_date = "07/15/2019"
    aug_ship_date = "08/15/2019"
    oct_ship_date = "10/15/2019"
    nov_ship_date = "11/15/2019"
    demo_ship_date = "11/15/2019"

    # set up variables based on order type.
    if (input_ordertype == '1'):
      # ski prsn order
      output_filetype = 'S'
      # ship dates
      col1_ship = aug_ship_date
      col2_ship = oct_ship_date
      col3_ship = nov_ship_date
      col4_ship = 'n/a'
      # line params
      max_line_number = 165
    elif (input_ordertype == '2'):
      # phantom order
      output_filetype = 'P'
      # ship dates
      col1_ship = aug_ship_date
      col2_ship = oct_ship_date
      col3_ship = nov_ship_date
      col4_ship = 'n/a'
      # line params
      max_line_number = 24
    elif (input_ordertype == '3'):
      # dreamtime order
      output_filetype = 'D'
      # ship dates
      col1_ship = dt_ship_date
      col2_ship = 'n/a'
      col3_ship = 'n/a'
      col4_ship = 'n/a'
      # line params
      max_line_number = 32
    else:
      # gotta be #4 demo/rental
      output_filetype = 'R'
      # ship dates
      col1_ship = 'n/a'
      col2_ship = 'n/a'
      col3_ship = demo_ship_date
      col4_ship = 'n/a'
      # line params
      max_line_number = 149

    # read in the spreadsheet
    wb = load_workbook("inputs/" + input_fn, data_only=True)

    # print(wb.sheetnames)

    # point to the proper sheet
    ws = wb['F19 order form']

    # Get Header data and stuff into program variables
    acct_name = ws['C7'].value
    buyer_name = ws['C8'].value
    buyer_email = ws['C9'].value
    customer_number = ws['J7'].value
    shp_addr_id = ws['J8'].value
    discount = ws['J9'].value
    order_note = ws['J10'].value
    col1_po = ws['E14'].value
    col2_po = ws['F14'].value
    col3_po = ws['G14'].value
    col4_po = ws['H14'].value

    # create an empty lines list for each ship date
    lines_august = []
    lines_october = []
    lines_november = []
    lines_demo = []

    # print('***** HEADER ******')
    # print(acct_name)
    # print(buyer_name)
    # print(buyer_email)
    # print(customer_number)
    # print(rep_name)
    # print(discount)
    # # print(po_number)

    # Line data
    # print('***** LINE DATA ******')
    # cycle through the appropriate rows... min and max rows will change with number of skus
    # and get line data
    # but first change a blank quantity field to 0
    for row in ws.iter_rows(min_row=16, max_col=13, max_row=max_line_number):
      if row[3].value != None: # process only if not a total row
        sku = row[3].value
        if row[11].value is None:
          row[11].value = 0
        elif row[11].value == 'N/C':
          row[11].value = 0          
        net_price = round(row[11].value, 2)
        if row[4].value is None:
          row[4].value = 0
        if row[5].value is None:
          row[5].value = 0
        if row[6].value is None:
          row[6].value = 0
        if row[7].value is None:
          row[7].value = 0
        total_sku_units = row[4].value + row[5].value + row[6].value + row[7].value
        # only need to process the line if there are some units ordered
        if total_sku_units > 0: # some units ordered for this sku so create order lines
          # print('****' + row[3].value + '****' + str(total_sku_units))
          # we know this sku was ordered, put the line in the proper ship date lines list
          if row[4].value > 0: # add line to Aug order
            lines_august.append({'sku': sku, 'quantity': row[4].value, 'net_price': net_price})
          if row[5].value > 0: # add line to Oct order
            lines_october.append({'sku': sku, 'quantity': row[5].value, 'net_price': net_price})
          if row[6].value > 0: # add line to Nov order
            lines_november.append({'sku': sku, 'quantity': row[6].value, 'net_price': net_price})
          if row[7].value > 0: # add line to demo order
            lines_demo.append({'sku': sku, 'quantity': row[7].value, 'net_price': net_price})

    # now build the orders[]

    orders = []

    # if lines_*** length is > 0 then build the month order
    if len(lines_august) > 0:
      orders.append({ 'customer_number': customer_number, 
                      'ship_address_id': shp_addr_id, 
                      'buyer_name': buyer_name,
                      'buyer_email': buyer_email,
                      'order_note': order_note,
                      'ship_date': col1_ship, 
                      'po_number': col1_po, 
                      'earned_discount': discount, 
                      'lines': lines_august })

    if len(lines_october) > 0:
      orders.append({ 'customer_number': customer_number, 
                      'ship_address_id': shp_addr_id, 
                      'buyer_name': buyer_name,
                      'buyer_email': buyer_email,
                      'order_note': order_note,
                      'ship_date': col2_ship, 
                      'po_number': col2_po, 
                      'earned_discount': discount, 
                      'lines': lines_october })

    if len(lines_november) > 0:
      orders.append({ 'customer_number': customer_number,
                      'ship_address_id': shp_addr_id, 
                      'buyer_name': buyer_name,
                      'buyer_email': buyer_email,
                      'order_note': order_note,
                      'ship_date': col3_ship, 
                      'po_number': col3_po, 
                      'earned_discount': discount, 
                      'lines': lines_november })

    if len(lines_demo) > 0:
      orders.append({ 'customer_number': customer_number, 
                      'ship_address_id': shp_addr_id, 
                      'buyer_name': buyer_name,
                      'buyer_email': buyer_email,
                      'order_note': order_note,
                      'ship_date': col4_ship, 
                      'po_number': col4_po, 
                      'earned_discount': discount, 
                      'lines': lines_demo })

    # magic that sets up the data element names correctly
    my_item_func = lambda x: x[:-1]

    # build the orders file 
    preseason_order = dicttoxml(orders, attr_type=False, custom_root='orders', item_func=my_item_func)

    print(parseString(preseason_order).toprettyxml())
    print('*****************************************')
    print("writing to file " + customer_number + "_" + shp_addr_id + "_" + output_filetype + ".xml in 'outputs' folder")
    print('*****************************************')

    # write output file
    output_file = open("outputs/" + customer_number + "_" + shp_addr_id + "_" + output_filetype + ".xml", "w")
    output_file.write(parseString(preseason_order).toprettyxml())
    output_file.close()

    print(" ")
    print("Finished Processing " + input_fn)
    print('*****************************************')
    print('  ')
    r_u_done = input("Hit ENTER to process another file or type 'end' and hit ENTER to quit:  ")
    if (r_u_done == 'end'):
      DONE = True

  else: # if input file not found
    print('*****************************************')
    print('  ')
    print('File "'+ input_fn + '" not found in "inputs" folder. Try again.')
    print('  ')

# that's all folks
print('  ')
print("Thanks for using the F19 DPS Order Wizard!")